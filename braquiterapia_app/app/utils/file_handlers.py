"""
Utilidades para manejo de archivos Excel y PDF
"""
import os
import io
import tempfile
import subprocess
import platform
import shutil
import fitz  # PyMuPDF
from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell


def get_libreoffice_path():
    """
    Encuentra la ruta de LibreOffice de manera multiplataforma.
    
    Returns:
        str: Ruta al ejecutable de LibreOffice
    
    Raises:
        RuntimeError: Si no se encuentra LibreOffice
    """
    # Intentar encontrar en PATH
    soffice = shutil.which('soffice')
    if soffice:
        return soffice
    
    # Rutas comunes por sistema operativo
    if platform.system() == 'Windows':
        paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"
        ]
    elif platform.system() == 'Darwin':  # macOS
        paths = ['/Applications/LibreOffice.app/Contents/MacOS/soffice']
    else:  # Linux
        paths = ['/usr/bin/soffice', '/usr/local/bin/soffice']
    
    for path in paths:
        if os.path.exists(path):
            return path
    
    raise RuntimeError(
        "LibreOffice no encontrado. Por favor instale LibreOffice desde https://www.libreoffice.org/"
    )


def xlsx_to_pdf(xlsx_bytes):
    """
    Convierte un archivo XLSX a PDF usando LibreOffice.
    
    Args:
        xlsx_bytes: BytesIO o bytes del archivo Excel
    
    Returns:
        BytesIO: Archivo PDF resultante
    
    Raises:
        RuntimeError: Si la conversión falla
    """
    # Envolver bytes si es necesario
    if isinstance(xlsx_bytes, (bytes, bytearray)):
        xlsx_bytes = io.BytesIO(xlsx_bytes)
    
    with tempfile.TemporaryDirectory() as tmpdir:
        # Guardar XLSX temporal
        xlsx_path = os.path.join(tmpdir, "temp.xlsx")
        xlsx_bytes.seek(0)
        with open(xlsx_path, "wb") as f:
            f.write(xlsx_bytes.getvalue())
        
        # Obtener ruta de LibreOffice
        soffice_path = get_libreoffice_path()
        
        # Comando de conversión
        cmd = [
            soffice_path,
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--convert-to", "pdf",
            "--outdir", tmpdir,
            xlsx_path,
        ]
        
        # Ejecutar conversión
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=30
        )
        
        # Buscar PDF generado
        pdfs = [f for f in os.listdir(tmpdir) if f.lower().endswith(".pdf")]
        
        if result.returncode != 0 or not pdfs:
            error_msg = result.stderr.decode(errors="ignore")
            raise RuntimeError(
                f"No se pudo convertir el XLSX a PDF.\n"
                f"Código de salida: {result.returncode}\n"
                f"Error: {error_msg}"
            )
        
        # Leer PDF generado
        pdf_path = os.path.join(tmpdir, pdfs[0])
        with open(pdf_path, "rb") as f:
            pdf_bytes = io.BytesIO(f.read())
        
        pdf_bytes.seek(0)
        return pdf_bytes


def pdf_to_png(pdf_bytes, rotation=90, dpi=150):
    """
    Convierte la primera página de un PDF a PNG.
    
    Args:
        pdf_bytes: Bytes del archivo PDF
        rotation: Grados de rotación (0, 90, 180, 270)
        dpi: Resolución de la imagen
    
    Returns:
        bytes: Imagen PNG
    """
    # Abrir PDF
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[0]
    
    # Renderizar a imagen
    pix = page.get_pixmap(dpi=dpi)
    png_bytes = pix.tobytes("png")
    doc.close()
    
    # Rotar usando PIL si es necesario
    if rotation != 0:
        img = Image.open(io.BytesIO(png_bytes))
        rotated = img.rotate(rotation, expand=True)
        
        out = io.BytesIO()
        rotated.save(out, format="PNG")
        png_bytes = out.getvalue()
    
    return png_bytes


def insert_png_into_excel(excel_bytes, png_bytes, sheet_name="IMAGEN", cell="B5", scale=0.35):
    """
    Inserta una imagen PNG en un archivo Excel.
    
    Args:
        excel_bytes: Bytes del archivo Excel
        png_bytes: Bytes de la imagen PNG
        sheet_name: Nombre de la hoja donde insertar
        cell: Celda donde anclar la imagen (ej: "B5")
        scale: Factor de escala de la imagen
    
    Returns:
        bytes: Excel modificado
    """
    # Cargar Excel
    wb = load_workbook(io.BytesIO(excel_bytes))
    
    # Obtener hoja (crear si no existe)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]
    
    # Crear objeto imagen de openpyxl
    img = XLImage(io.BytesIO(png_bytes))
    
    # Escalar imagen
    img.width = img.width * scale
    img.height = img.height * scale
    
    # Insertar en celda
    ws.add_image(img, cell)
    
    # Guardar a bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def write_to_excel_cell(wb, sheet_name, cell, value, alignment='center'):
    """
    Escribe un valor en una celda de Excel, manejando celdas combinadas.
    
    Args:
        wb: Workbook de openpyxl
        sheet_name: Nombre de la hoja
        cell: Celda destino (ej: "C8")
        value: Valor a escribir
        alignment: Alineación ('left', 'center', 'right')
    
    Returns:
        bool: True si tuvo éxito
    """
    if sheet_name not in wb.sheetnames:
        return False
    
    ws = wb[sheet_name]
    cell_obj = ws[cell]
    
    # Manejar celdas combinadas
    if isinstance(cell_obj, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell in str(merged_range):
                top_left = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                )
                top_left.value = value
                cell_obj = top_left
                break
    else:
        cell_obj.value = value
    
    # Aplicar alineación
    if alignment:
        align_map = {
            'left': Alignment(horizontal='left'),
            'center': Alignment(horizontal='center'),
            'right': Alignment(horizontal='right')
        }
        cell_obj.alignment = align_map.get(alignment, Alignment(horizontal='center'))
    
    return True
