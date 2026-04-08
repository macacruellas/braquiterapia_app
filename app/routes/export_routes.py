"""
Rutas para exportación de informes (Cartón dosimétrico e Informe final)
"""
import os
import io
import json
from datetime import datetime
from flask import Blueprint, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from app.utils.file_handlers import (
    xlsx_to_pdf,
    pdf_to_png,
    insert_png_into_excel,
    write_to_excel_cell
)
from app.utils.helpers import parse_patient_name, round_2_decimals
from config.settings import TEMPLATE_CARTON, TEMPLATE_INFORME

bp = Blueprint('export', __name__)

_MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]

def _format_fechas_es(date_strings):
    """
    Recibe una lista de strings 'YYYY-MM-DD' y devuelve una cadena en español.
    Ej: ['2020-10-20','2020-10-22','2020-10-27','2020-10-29']
        -> '20, 22, 27 y 29 de Octubre de 2020'
    Si los días abarcan distintos meses:
        -> '20 de Octubre y 3 de Noviembre de 2020'
    """
    parsed = []
    for s in date_strings:
        s = (s or "").strip()
        if not s:
            continue
        try:
            parsed.append(datetime.strptime(s, "%Y-%m-%d"))
        except ValueError:
            pass

    if not parsed:
        return ""

    parsed.sort()

    mismo_mes_anio = all(
        d.month == parsed[0].month and d.year == parsed[0].year
        for d in parsed
    )

    if mismo_mes_anio:
        mes  = _MESES_ES[parsed[0].month - 1].capitalize()
        anio = parsed[0].year
        dias = [str(d.day) for d in parsed]
        if len(dias) == 1:
            return f"{dias[0]} de {mes} de {anio}"
        return f"{', '.join(dias[:-1])} y {dias[-1]} de {mes} de {anio}"
    else:
        partes = [
            f"{d.day} de {_MESES_ES[d.month - 1].capitalize()} de {d.year}"
            for d in parsed
        ]
        if len(partes) == 1:
            return partes[0]
        return f"{', '.join(partes[:-1])} y {partes[-1]}"


@bp.route("/export_carton", methods=["POST"])
def export_carton():
    """
    Exporta el cartón dosimétrico en formato PDF.
    Incluye datos de RT externa y braquiterapia en formato oficial.
    """
    # Recuperar datos desde el payload
    payload = request.form.get("payload", "")
    if not payload:
        return "Sin datos para exportar", 400
    
    try:
        data = json.loads(payload)
    except Exception as e:
        return f"Payload inválido: {str(e)}", 400
    
    # Extraer datos
    patient_name = data.get("patient_name") or ""
    patient_id = data.get("patient_id") or ""
    fx_rt = int(data.get("fx_rt") or 0)
    n_hdr = int(data.get("n_hdr") or 0)
    summary = data.get("summary") or []
    ebrt = data.get("ebrt") or []
    hdr_fractions = data.get("hdr_fractions") or []

    
    # Verificar que existe la plantilla
    if not os.path.exists(TEMPLATE_CARTON):
    # Debug: mostrar la ruta completa que está buscando
        import sys
        return f"""
    <h2>Error: Plantilla no encontrada</h2>
    <p><strong>Ruta buscada:</strong> {TEMPLATE_CARTON}</p>
    <p><strong>¿Existe el archivo?:</strong> {os.path.exists(TEMPLATE_CARTON)}</p>
    <p><strong>Directorio de trabajo actual:</strong> {os.getcwd()}</p>
    <p>Por favor, asegurate de que el archivo "Cartón dosimétrico.xlsx" esté en la carpeta app/templates/</p>
    """, 500
    
    # Abrir plantilla
    wb = load_workbook(TEMPLATE_CARTON)
    ws = wb["Hoja1 (2)"] if "Hoja1 (2)" in wb.sheetnames else wb.active
    
    # Separar apellido y nombre
    apellido, nombre = parse_patient_name(patient_name)
    
    # Escribir datos del paciente
    write_to_excel_cell(wb, ws.title, "C8", (apellido or patient_name).upper(), 'left')
    write_to_excel_cell(wb, ws.title, "C9", (nombre or "").upper(), 'left')
    write_to_excel_cell(wb, ws.title, "H3", patient_id, 'center')
    
    # === Tratamiento de RT Externa ===
    # Dosis total = fracciones × 2 Gy
    write_to_excel_cell(wb, ws.title, "C12", round(fx_rt * 2, 2), 'center')
    write_to_excel_cell(wb, ws.title, "C13", fx_rt, 'center')
    
    # Mapa de EBRT
    ebrt_map = {}
    for row in ebrt:
        roi = (row.get("roi") or "").upper()
        if roi:
            ebrt_map[roi] = row
    
    # CTV D95
    ctv_row = ebrt_map.get("CTV")
    if ctv_row:
        dose_ctv = round_2_decimals(ctv_row.get("D_ext"))
        write_to_excel_cell(wb, ws.title, "C14", dose_ctv, 'center')
    
    # OARs D2cc
    oar_rows = [
        ("RECTO", 13),
        ("VEJIGA", 14),
        ("SIGMOIDE", 15),
        ("INTESTINO", 16),
    ]
    for roi_key, row_idx in oar_rows:
        row = ebrt_map.get(roi_key)
        if row:
            dose_val = round_2_decimals(row.get("D_ext"))
            ws.cell(row=row_idx, column=9).value = dose_val
            ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center')
    
    # === Tabla de Tratamiento HDR (EQD2 por sesión) ===
    row_map_hdr = {
        "CTV": 24,
        "Recto": 25,
        "Vejiga": 26,
        "Sigmoide": 27,
        "Intestino": 28,
    }
    
    hdr_map = {(x["roi"] or "").upper(): x for x in hdr_fractions}
    
    def match_roi(roi_excel):
        roi_excel = roi_excel.upper()
        for roi_hdr, data in hdr_map.items():
            if roi_excel == "CTV" and "CTV" in roi_hdr:
                return data
            if roi_excel != "CTV" and roi_excel in roi_hdr:
                return data
        return None
    
    # Columnas de sesiones: C, D, F, H
    session_cols = [3, 4, 6, 8]
    
    for roi_excel, row_idx in row_map_hdr.items():
        item = match_roi(roi_excel)
        doses = [round_2_decimals(v) for v in item["doses"]] if item else []
        for j in range(4):  # 4 sesiones máximo
            col = session_cols[j]
            cell = ws.cell(row=row_idx, column=col)
            
            # Manejar celdas combinadas
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        cell = ws.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        break
            
            if j < len(doses) and doses[j] is not None:
                cell.value = doses[j]
            else:
                cell.value = "-"
            
            cell.alignment = Alignment(horizontal='center')
    
    # === Registro de dosis total (EQD2) ===
    row_map_total = {
        "CTV": 35,
        "RECTO": 36,
        "VEJIGA": 37,
        "SIGMOIDE": 38,
        "INTESTINO": 39,
    }
    
    for item in summary:
        roi_raw = (item.get("roi") or "").upper()
        roi_key = roi_raw.replace(" (D90)", "")
        row_idx = row_map_total.get(roi_key)
        if not row_idx:
            continue
        
        # EQD2 RT Externa (columna C)
        ws.cell(row=row_idx, column=3).value = round_2_decimals(item.get("eqd2_ebrt"))
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')
        
        # EQD2 HDR (columna D)
        ws.cell(row=row_idx, column=4).value = round_2_decimals(item.get("eqd2_hdr"))
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
        
        # EQD2 TOTAL (columna G)
        ws.cell(row=row_idx, column=7).value = round_2_decimals(item.get("eqd2_total"))
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
    
    # Guardar Excel en memoria
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_bytes = excel_buffer.getvalue()
    
    # Convertir a PDF
    try:
        pdf_bytes = xlsx_to_pdf(excel_bytes)
    except Exception as e:
        return f"Error al convertir a PDF: {str(e)}", 500
    
    # Nombre del archivo
    filename = f"Carton_{(patient_id or patient_name or 'paciente').replace(' ', '_')}.pdf"
    
    return send_file(
        pdf_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )


@bp.route("/export_informe", methods=["POST"])
def export_informe():
    """
    Exporta el informe final en formato Excel editable.
    Opcionalmente incluye imagen del plan en PDF si se adjunta.
    """
    # Recuperar datos desde el payload
    payload = request.form.get("payload", "")
    if not payload:
        return "Sin datos para exportar", 400
    
    try:
        data = json.loads(payload)
    except Exception as e:
        return f"Payload inválido: {str(e)}", 400
    
    patient_name = data.get("patient_name") or ""
    patient_id = data.get("patient_id") or ""
    summary = data.get("summary") or []
    
    # Verificar que existe la plantilla
    if not os.path.exists(TEMPLATE_INFORME):
        return f"Plantilla no encontrada: {TEMPLATE_INFORME}", 500
    
    # Abrir plantilla
    wb = load_workbook(TEMPLATE_INFORME)
    ws = wb.worksheets[0]
    
    # Fecha actual
    write_to_excel_cell(wb, ws.title, "G7", datetime.today().strftime("%d/%m/%Y"))
    
    # Nombre del paciente
    write_to_excel_cell(wb, ws.title, "G12", patient_name)
    
    # Tabla de dosis (filas 32-36)
    row_map = {
        "CTV": 32,
        "RECTO": 33,
        "VEJIGA": 34,
        "SIGMOIDE": 35,
        "INTESTINO": 36,
    }
    
    for item in summary:
        roi_raw = (item.get("roi") or "").upper()
        roi_key = roi_raw.replace(" (D90)", "")
        row = row_map.get(roi_key)
        if not row:
            continue
        
        # Columna D: EQD2 RT Externa
        ws[f"D{row}"].value = round_2_decimals(item.get("eqd2_ebrt"))
        
        # Columna F: EQD2 HDR
        ws[f"F{row}"].value = round_2_decimals(item.get("eqd2_hdr"))
        
        # Columna H: EQD2 TOTAL
        ws[f"H{row}"].value = round_2_decimals(item.get("eqd2_total"))
    
    # === Campos adicionales del informe ===
    inf_diagnostico = request.form.get("inf_diagnostico", "").strip()
    inf_braqui     = request.form.get("inf_braqui", "").strip()
    inf_aplicador  = request.form.get("inf_aplicador", "").strip()
    inf_sesiones   = request.form.get("inf_sesiones", "").strip()
    inf_dosis_gy   = request.form.get("inf_dosis_gy", "").strip()

    if inf_diagnostico:
        write_to_excel_cell(wb, ws.title, "E13", inf_diagnostico, 'left')

    if inf_braqui:
        write_to_excel_cell(wb, ws.title, "G15", inf_braqui, 'left')

    if inf_aplicador:
        write_to_excel_cell(wb, ws.title, "C21", inf_aplicador, 'left')

    if inf_sesiones and inf_dosis_gy:
        try:
            n_ses = int(inf_sesiones)
            d_gy  = float(inf_dosis_gy)
            dosis_str = f"{n_ses} sesiones de {d_gy:g} Gy"
            write_to_excel_cell(wb, ws.title, "D24", dosis_str, 'center')
        except ValueError:
            pass

    # Fechas de tratamiento (E26)
    raw_fechas = [request.form.get(f"inf_fecha_{i}", "") for i in range(1, 5)]
    fechas_str = _format_fechas_es(raw_fechas)
    if fechas_str:
        write_to_excel_cell(wb, ws.title, "E26", fechas_str, 'left')

    # Duración del tratamiento (E27)
    inf_dur_num  = request.form.get("inf_dur_num", "").strip()
    inf_dur_unit = request.form.get("inf_dur_unit", "semanas").strip()
    if inf_dur_num:
        try:
            n_dur = int(inf_dur_num)
            write_to_excel_cell(wb, ws.title, "E27", f"{n_dur} {inf_dur_unit}", 'left')
        except ValueError:
            pass

    # Guardar Excel inicial en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    
    # === NUEVO: Si se subió un PDF, convertirlo a PNG e insertarlo ===
    pdf_file = request.files.get("plan_pdf")
    if pdf_file and pdf_file.filename:
        try:
            pdf_bytes = pdf_file.read()
            png_bytes = pdf_to_png(pdf_bytes, rotation=90, dpi=150)
            excel_bytes = insert_png_into_excel(
                excel_bytes,
                png_bytes,
                sheet_name="IMAGEN",
                cell="B5",
                scale=0.35
            )
        except Exception as e:
            # Si falla, continuar sin la imagen
            print(f"Advertencia: No se pudo insertar imagen del plan: {str(e)}")
    
    # Nombre del archivo
    filename = f"Informe_final_{(patient_id or patient_name or 'paciente').replace(' ', '_')}.xlsx"
    
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
